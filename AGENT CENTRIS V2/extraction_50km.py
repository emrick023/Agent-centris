#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Extraction Centris — Plex à vendre — 50km autour de Sherbrooke
Produit : plex_sherbrooke_50km.xlsx  +  envoie le fichier par email

Dépendances : openpyxl  →  pip install openpyxl
Variables d'environnement requises pour l'email :
  GMAIL_USER         ex: moncompte@gmail.com
  GMAIL_APP_PASSWORD ex: abcd efgh ijkl mnop  (mot de passe d'application Google)
  EMAIL_DEST         ex: destinataire@email.com

Usage local  : python3 extraction_50km.py
Usage GitHub Actions : le workflow injecte automatiquement les variables.
"""

import urllib.request
import urllib.error
import re
import time
import os
import sys
import html as html_module
import smtplib
import datetime
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("ERREUR : openpyxl n'est pas installé.")
    print("Installez-le avec :  pip install openpyxl")
    sys.exit(1)

# ---------------------------------------------------------------------------
# CONFIGURATION
# ---------------------------------------------------------------------------

# Chemin de sortie — relatif au répertoire du script (fonctionne local + GitHub Actions)
OUTPUT_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                           "plex_sherbrooke_50km.xlsx")

# Variables d'environnement pour l'email (injectées par GitHub Actions Secrets)
GMAIL_USER         = os.environ.get("GMAIL_USER", "")
GMAIL_APP_PASSWORD = os.environ.get("GMAIL_APP_PASSWORD", "")
EMAIL_DEST         = os.environ.get("EMAIL_DEST", "")

# Villes Centris dans un rayon de ~50 km autour de Sherbrooke
# Format : (nom_ville_affiche, slug_url_centris)
VILLES = [
    ("Sherbrooke",                        "sherbrooke"),
    ("Magog",                             "magog"),
    ("Coaticook",                         "coaticook"),
    ("Windsor",                           "windsor"),
    ("Waterville",                        "waterville"),
    ("Compton",                           "compton"),
    ("East Angus",                        "east-angus"),
    ("Val-des-Sources",                   "val-des-sources"),
    ("Saint-Denis-de-Brompton",           "saint-denis-de-brompton"),
    ("Lac-Mégantic",                      "lac-megantic"),
    ("Weedon",                            "weedon"),
    ("Cookshire-Eaton",                   "cookshire-eaton"),
    ("Danville",                          "danville"),
    ("Richmond",                          "richmond"),
    ("Melbourne",                         "melbourne"),
    ("Sainte-Catherine-de-Hatley",        "sainte-catherine-de-hatley"),
    ("North Hatley",                      "north-hatley"),
    ("Ayer's Cliff",                      "ayers-cliff"),
    ("Stanstead",                         "stanstead"),
    ("Stoke",                             "stoke"),
    ("Saint-François-Xavier-de-Brompton", "saint-francois-xavier-de-brompton"),
    ("Bromptonville",                     "bromptonville"),
    ("Asbestos",                          "asbestos"),
    ("Lawrenceville",                     "lawrenceville"),
    ("Barnston-Ouest",                    "barnston-ouest"),
    ("Hatley",                            "hatley"),
]

BASE_URL      = "https://www.centris.ca"
PAGE_SIZE     = 20      # Centris affiche 20 annonces par page
DELAY_PAGES   = 1.5     # secondes entre les pages de résultats
DELAY_FICHES  = 1.0     # secondes entre les fiches individuelles

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/122.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "fr-CA,fr;q=0.9,en;q=0.8",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
}

# ---------------------------------------------------------------------------
# HTTP
# ---------------------------------------------------------------------------

def fetch(url, retries=3):
    """Télécharge une URL et retourne le HTML en str (utf-8)."""
    for attempt in range(retries):
        try:
            req = urllib.request.Request(url, headers=HEADERS)
            with urllib.request.urlopen(req, timeout=25) as resp:
                raw = resp.read()
                try:
                    return raw.decode("utf-8")
                except UnicodeDecodeError:
                    return raw.decode("latin-1", errors="replace")
        except urllib.error.HTTPError as e:
            if e.code == 404:
                return ""   # ville inexistante sur Centris → on passe
            print(f"    [HTTP {e.code}] {url}")
            if attempt < retries - 1:
                time.sleep(2)
        except Exception as e:
            print(f"    [ERREUR] {url} → {e}")
            if attempt < retries - 1:
                time.sleep(2)
    return ""

# ---------------------------------------------------------------------------
# COLLECTE DES LIENS D'ANNONCES
# ---------------------------------------------------------------------------

LISTING_RE = re.compile(
    r'href="(/fr/(?:duplex|triplex|quadruplex|quintuplex|plex)~a-vendre~[^"/]+/(\d{7,}))"',
    re.IGNORECASE
)

NB_RESULTS_RE = re.compile(r'<span\s+id="numberOfResults"\s*>(\d+)<', re.IGNORECASE)


def get_listing_urls_for_ville(ville_nom, ville_slug):
    """Retourne un dict {listing_id: url} pour toutes les annonces d'une ville."""
    results = {}
    base = f"{BASE_URL}/fr/plex~a-vendre~{ville_slug}"

    html = fetch(base)
    if not html:
        print("    → Aucune réponse (ville absente de Centris ou erreur réseau)")
        return results

    nb_m  = NB_RESULTS_RE.search(html)
    total = int(nb_m.group(1)) if nb_m else 0
    nb_pages = max(1, (total + PAGE_SIZE - 1) // PAGE_SIZE)
    print(f"    Total annonces : {total}  ({nb_pages} page{'s' if nb_pages > 1 else ''})")

    for m in LISTING_RE.finditer(html):
        lid = m.group(2)
        if lid not in results:
            results[lid] = BASE_URL + m.group(1)

    for page in range(2, nb_pages + 1):
        url = f"{base}?page={page}"
        print(f"    Page {page}/{nb_pages} : {url}")
        time.sleep(DELAY_PAGES)
        html = fetch(url)
        if not html:
            break
        for m in LISTING_RE.finditer(html):
            lid = m.group(2)
            if lid not in results:
                results[lid] = BASE_URL + m.group(1)

    return results

# ---------------------------------------------------------------------------
# EXTRACTION D'UNE FICHE
# ---------------------------------------------------------------------------

def decode_html(text):
    """Décode les entités HTML et nettoie les espaces."""
    text = html_module.unescape(text)
    return re.sub(r'\s+', ' ', text).strip()


def clean_number(text):
    """Extrait un entier brut. Retourne 'Non indiqué' si absent."""
    if not text:
        return "Non indiqué"
    digits = re.sub(r'[^\d]', '', str(text))
    return int(digits) if digits else "Non indiqué"


def carac_value(html, label):
    """
    Cherche un bloc carac-container dont le carac-title contient `label`,
    retourne le texte du carac-value associé.
    """
    pat = re.compile(
        re.escape(label) + r'.*?</div>\s*<div class="carac-value"[^>]*>'
        r'<span[^>]*>(.*?)</span>',
        re.DOTALL | re.IGNORECASE
    )
    m = pat.search(html)
    return decode_html(m.group(1)) if m else "Non indiqué"


def financial_total(html, section_label):
    """
    Retourne le Total annuel de la section financière dont le titre contient
    `section_label` (ex: "Taxes", "valuation municipale", "penses").
    """
    sec_pat = re.compile(
        section_label + r'.*?financial-details-table-total[^>]*>'
        r'.*?<td[^>]*class="font-weight-bold text-right"[^>]*>([\d\xA0\s,]+\$?)</td>',
        re.DOTALL | re.IGNORECASE
    )
    m = sec_pat.search(html)
    if m:
        return clean_number(m.group(1))

    # Fallback : premier total après le label
    label_idx = html.lower().find(section_label.lower())
    if label_idx == -1:
        return "Non indiqué"
    chunk = html[label_idx: label_idx + 4000]
    total_m = re.search(
        r'font-weight-bold text-right"[^>]*>([\d\xA0\s,]+\$?)</td>',
        chunk, re.IGNORECASE
    )
    return clean_number(total_m.group(1)) if total_m else "Non indiqué"


def extract_listing(url, ville_nom):
    """Extrait toutes les données d'une fiche Centris."""
    html = fetch(url)
    if not html:
        return None

    data = {"Ville": ville_nom, "URL": url}

    # Adresse
    addr_m = re.search(r'<h2\s+itemprop="address"[^>]*>\s*(.*?)\s*</h2>',
                       html, re.DOTALL | re.IGNORECASE)
    data["Adresse"] = decode_html(addr_m.group(1)) if addr_m else "Non indiqué"

    # Prix
    prix_m = re.search(r'<meta\s+itemprop="price"\s+content="(\d+)"', html, re.IGNORECASE)
    data["Prix"] = int(prix_m.group(1)) if prix_m else "Non indiqué"

    # Nombre d'unités → extraire le chiffre de "Résidentiel (N)"
    nb_raw = carac_value(html, "Nombre d")
    if nb_raw != "Non indiqué":
        nb_m = re.search(r'(\d+)', nb_raw)
        data["Nombre d'unités"] = int(nb_m.group(1)) if nb_m else "Non indiqué"
    else:
        data["Nombre d'unités"] = "Non indiqué"

    # Unités résidentielles (ex: "2 x 3 ½")
    data["Unités résidentielles"] = carac_value(html, "nités résidentielles")

    # Année de construction
    annee_raw = carac_value(html, "de construction")
    if annee_raw != "Non indiqué":
        annee_m = re.search(r'(\d{4})', annee_raw)
        yr = int(annee_m.group(1)) if annee_m else None
        data["Année de construction"] = yr if yr and 1800 <= yr <= 2030 else "Non indiqué"
    else:
        data["Année de construction"] = "Non indiqué"

    # Superficie du terrain
    surf_raw = carac_value(html, "Superficie du terrain")
    data["Superficie terrain"] = clean_number(surf_raw)

    # Revenu brut potentiel
    rev_raw = carac_value(html, "Revenus bruts potentiels")
    data["Revenu brut potentiel"] = clean_number(rev_raw)

    # Évaluation municipale
    data["Évaluation municipale"] = financial_total(html, "valuation municipale")

    # Taxes
    data["Taxes"] = financial_total(html, "Taxes")

    # Dépenses
    data["Dépenses"] = financial_total(html, "penses")

    return data

# ---------------------------------------------------------------------------
# EXCEL
# ---------------------------------------------------------------------------

COLONNES = [
    "Ville",
    "Adresse",
    "Prix",
    "Nombre d'unités",
    "Année de construction",
    "Superficie terrain",
    "Unités résidentielles",
    "Revenu brut potentiel",
    "Évaluation municipale",
    "Taxes",
    "Dépenses",
    "URL",
]

COL_WIDTHS = {
    "Ville": 24,
    "Adresse": 46,
    "Prix": 14,
    "Nombre d'unités": 16,
    "Année de construction": 20,
    "Superficie terrain": 18,
    "Unités résidentielles": 32,
    "Revenu brut potentiel": 22,
    "Évaluation municipale": 22,
    "Taxes": 12,
    "Dépenses": 14,
    "URL": 65,
}


def save_excel(rows, path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Annonces admissibles"

    h_font  = Font(bold=True, color="FFFFFF")
    h_fill  = PatternFill("solid", fgColor="1F4E79")
    h_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, col_name in enumerate(COLONNES, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font      = h_font
        cell.fill      = h_fill
        cell.alignment = h_align

    ws.row_dimensions[1].height = 30

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, col_name in enumerate(COLONNES, start=1):
            val  = row.get(col_name, "Non indiqué")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = Alignment(vertical="center")

    for col_idx, col_name in enumerate(COLONNES, start=1):
        letter = openpyxl.utils.get_column_letter(col_idx)
        ws.column_dimensions[letter].width = COL_WIDTHS.get(col_name, 20)

    ws.freeze_panes = "A2"

    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb.save(path)
    print(f"Fichier Excel sauvegardé : {path}")

# ---------------------------------------------------------------------------
# EMAIL
# ---------------------------------------------------------------------------

def send_email(excel_path, nb_annonces):
    """Envoie le fichier Excel par email via Gmail SMTP (App Password)."""
    if not GMAIL_USER or not GMAIL_APP_PASSWORD or not EMAIL_DEST:
        print("[EMAIL] Variables d'environnement manquantes — email non envoyé.")
        print("        Définir : GMAIL_USER, GMAIL_APP_PASSWORD, EMAIL_DEST")
        return

    today   = datetime.date.today().strftime("%d %B %Y")
    subject = f"Centris — Plex 50km Sherbrooke — {today} ({nb_annonces} annonces)"

    body = f"""Bonjour,

Voici le rapport quotidien des plex à vendre dans un rayon de 50 km autour de Sherbrooke.

Date       : {today}
Annonces   : {nb_annonces}
Villes     : {len(VILLES)} municipalités couvertes

Le fichier Excel est joint à cet email.

---
Généré automatiquement par le script Centris.
"""

    msg = MIMEMultipart()
    msg["From"]    = GMAIL_USER
    msg["To"]      = EMAIL_DEST
    msg["Subject"] = subject
    msg.attach(MIMEText(body, "plain", "utf-8"))

    # Pièce jointe Excel
    filename = os.path.basename(excel_path)
    with open(excel_path, "rb") as f:
        part = MIMEBase("application",
                        "vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        part.set_payload(f.read())
    encoders.encode_base64(part)
    part.add_header("Content-Disposition", f'attachment; filename="{filename}"')
    msg.attach(part)

    try:
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
            server.login(GMAIL_USER, GMAIL_APP_PASSWORD)
            server.sendmail(GMAIL_USER, EMAIL_DEST, msg.as_string())
        print(f"[EMAIL] Envoyé avec succès à {EMAIL_DEST}")
    except Exception as e:
        print(f"[EMAIL] Erreur lors de l'envoi : {e}")

# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------

def main():
    print("=" * 65)
    print("EXTRACTION CENTRIS — PLEX À VENDRE — ~50 KM AUTOUR DE SHERBROOKE")
    print("=" * 65)

    # Étape 1 : collecter tous les liens
    print("\n[ÉTAPE 1] Collecte des liens d'annonces par ville...")
    all_listings = {}   # listing_id → (url, ville_nom)

    for ville_nom, ville_slug in VILLES:
        print(f"\n  {ville_nom}")
        ville_res = get_listing_urls_for_ville(ville_nom, ville_slug)
        for lid, url in ville_res.items():
            if lid not in all_listings:
                all_listings[lid] = (url, ville_nom)
        time.sleep(DELAY_PAGES)

    total = len(all_listings)
    print(f"\n  Total annonces uniques : {total}")

    if total == 0:
        print("\nAucune annonce trouvée. Vérifiez la connectivité.")
        return

    # Étape 2 : extraire chaque fiche
    print("\n[ÉTAPE 2] Extraction des données de chaque fiche...")
    rows = []

    for idx, (lid, (url, ville_nom)) in enumerate(all_listings.items(), start=1):
        print(f"  [{idx:3d}/{total}] {ville_nom} — {url}")
        data = extract_listing(url, ville_nom)
        if data:
            rows.append(data)
        time.sleep(DELAY_FICHES)

    print(f"\n  Fiches extraites avec succès : {len(rows)}")

    # Étape 3 : générer l'Excel
    print("\n[ÉTAPE 3] Génération du fichier Excel...")
    save_excel(rows, OUTPUT_PATH)

    # Étape 4 : envoyer par email
    print("\n[ÉTAPE 4] Envoi de l'email...")
    send_email(OUTPUT_PATH, len(rows))

    print("\nTerminé.")


if __name__ == "__main__":
    main()
