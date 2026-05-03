#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Extraction Centris — Plex à vendre — 50km autour de Sherbrooke
Produit : plex_sherbrooke_50km.xlsx  +  envoie le fichier par email SI changements détectés

Dépendances : openpyxl  →  pip install openpyxl
Variables d'environnement requises pour l'email :
  GMAIL_USER         ex: moncompte@gmail.com
  GMAIL_APP_PASSWORD ex: abcd efgh ijkl mnop  (mot de passe d'application Google)
  EMAIL_DEST         ex: destinataire@email.com

Usage local  : python3 extraction_50km.py
Usage GitHub Actions : le workflow injecte automatiquement les variables.
"""

import urllib.request
import urllib.error
import urllib.parse
import re
import time
import os
import sys
import json
import html as html_module
import smtplib
import datetime
import unicodedata
import argparse
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("ERREUR : openpyxl n'est pas installé.")
    print("Installez-le avec :  pip install openpyxl")
    sys.exit(1)

# ---------------------------------------------------------------------------
# CONFIGURATION
# ---------------------------------------------------------------------------

OUTPUT_PATH    = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                              "plex_sherbrooke_50km.xlsx")
REFERENCE_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                              "listing_ids_precedents.json")

GMAIL_USER         = os.environ.get("GMAIL_USER", "")
GMAIL_APP_PASSWORD = os.environ.get("GMAIL_APP_PASSWORD", "")
EMAIL_DEST         = os.environ.get("EMAIL_DEST", "")
RESET_REFERENCE    = os.environ.get("RESET_REFERENCE", "").lower() in {"1", "true", "yes"}

VILLES = [
    ("Sherbrooke",                        "sherbrooke"),
    ("Magog",                             "magog"),
    ("Coaticook",                         "coaticook"),
    ("Windsor",                           "windsor"),
    ("Waterville",                        "waterville"),
    ("Compton",                           "compton"),
    ("East Angus",                        "east-angus"),
    ("Val-des-Sources",                   "val-des-sources"),
    ("Saint-Denis-de-Brompton",           "saint-denis-de-brompton"),
    ("Lac-Mégantic",                      "lac-megantic"),
    ("Weedon",                            "weedon"),
    ("Cookshire-Eaton",                   "cookshire-eaton"),
    ("Danville",                          "danville"),
    ("Richmond",                          "richmond"),
    ("Melbourne",                         "melbourne"),
    ("Sainte-Catherine-de-Hatley",        "sainte-catherine-de-hatley"),
    ("North Hatley",                      "north-hatley"),
    ("Ayer's Cliff",                      "ayers-cliff"),
    ("Stanstead",                         "stanstead"),
    ("Stoke",                             "stoke"),
    ("Saint-François-Xavier-de-Brompton", "saint-francois-xavier-de-brompton"),
    ("Bromptonville",                     "bromptonville"),
    ("Asbestos",                          "asbestos"),
    ("Lawrenceville",                     "lawrenceville"),
    ("Barnston-Ouest",                    "barnston-ouest"),
    ("Hatley",                            "hatley"),
]

BASE_URL      = "https://www.centris.ca"
PAGE_SIZE     = 20
DELAY_PAGES   = 1.5
DELAY_FICHES  = 1.0

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/122.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "fr-CA,fr;q=0.9,en;q=0.8",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
}

# Couleurs highlighting Excel
CLR_NOUVEAU     = "C6EFCE"   # vert  — nouvelle annonce
CLR_PRIX_LIGNE  = "FFEB9C"   # jaune — changement de prix (ligne)
CLR_PRIX_CELL   = "FF6600"   # orange — changement de prix (cellule Prix)

# Garde-fous contre les extractions partielles: si Centris ou le réseau
# retourne une fraction inhabituelle des annonces, on évite de polluer la
# référence de demain et d'envoyer un rapport trompeur.
MAX_REMOVED_RATIO_FOR_SAFE_RUN = 0.10
MIN_ACTIVE_RATIO_FOR_SAFE_RUN  = 0.90
MAX_REMOVED_COUNT_FOR_SAFE_RUN = 10

# ---------------------------------------------------------------------------
# HTTP
# ---------------------------------------------------------------------------

def fetch_response(url, retries=3):
    for attempt in range(retries):
        try:
            req = urllib.request.Request(url, headers=HEADERS)
            with urllib.request.urlopen(req, timeout=25) as resp:
                raw = resp.read()
                try:
                    return raw.decode("utf-8"), resp.status
                except UnicodeDecodeError:
                    return raw.decode("latin-1", errors="replace"), resp.status
        except urllib.error.HTTPError as e:
            if e.code == 404:
                return "", 404
            print(f"    [HTTP {e.code}] {url}")
            if attempt < retries - 1:
                time.sleep(2)
        except Exception as e:
            print(f"    [ERREUR] {url} → {e}")
            if attempt < retries - 1:
                time.sleep(2)
    return "", None


def fetch(url, retries=3):
    html, _status = fetch_response(url, retries)
    return html

# ---------------------------------------------------------------------------
# COLLECTE DES LIENS
# ---------------------------------------------------------------------------

LISTING_RE    = re.compile(
    r'href="((?:https://www\.centris\.ca)?/fr/(?:duplex|triplex|quadruplex|quintuplex|plex)~a-vendre~[^"/]+/(\d{7,}))"',
    re.IGNORECASE
)
NB_RESULTS_RE = re.compile(r'<span\s+id="numberOfResults"\s*>(\d+)<', re.IGNORECASE)


def absolute_listing_url(href):
    return href if href.startswith("http") else BASE_URL + href


def hidden_span_value(html, span_id):
    pat = re.compile(
        r'<span\s+id="' + re.escape(span_id) + r'"\s*>(.*?)</span>',
        re.DOTALL | re.IGNORECASE
    )
    m = pat.search(html)
    return html_module.unescape(m.group(1)).strip() if m else ""


def build_stable_page_url(base, first_page_html, page):
    """
    Centris shuffles result pages unless the run's sortSeed is reused.
    Keeping the seed prevents overlapping pages and false "new" listings.
    """
    current_sort = hidden_span_value(first_page_html, "currentSort")
    if not current_sort or current_sort.lower() == "none":
        current_sort = "DateDesc"

    params = {
        "sort": current_sort,
        "sortSeed": hidden_span_value(first_page_html, "sortSeed"),
        "pageSize": hidden_span_value(first_page_html, "pageSize") or str(PAGE_SIZE),
        "q": hidden_span_value(first_page_html, "serializedSearchQuery"),
        "page": str(page),
    }
    params = {key: val for key, val in params.items() if val}

    if not params.get("sortSeed"):
        return f"{base}?page={page}"

    return f"{base}?{urllib.parse.urlencode(params)}"


def get_listing_urls_for_ville(ville_nom, ville_slug, return_stats=False):
    results = {}
    base    = f"{BASE_URL}/fr/plex~a-vendre~{ville_slug}"
    stats   = {"expected": 0, "collected": 0, "complete": True}

    html, status = fetch_response(base)
    if not html:
        if status == 404:
            print("    → Ville absente de Centris (0 annonce)")
        else:
            stats["complete"] = False
            print("    → Aucune réponse (erreur réseau ou Centris indisponible)")
        return (results, stats) if return_stats else results

    nb_m     = NB_RESULTS_RE.search(html)
    total    = int(nb_m.group(1)) if nb_m else 0
    stats["expected"] = total
    nb_pages = max(1, (total + PAGE_SIZE - 1) // PAGE_SIZE)
    print(f"    Total annonces : {total}  ({nb_pages} page{'s' if nb_pages > 1 else ''})")

    for m in LISTING_RE.finditer(html):
        lid = m.group(2)
        if lid not in results:
            results[lid] = absolute_listing_url(m.group(1))

    first_page_html = html

    for page in range(2, nb_pages + 1):
        url = build_stable_page_url(base, first_page_html, page)
        print(f"    Page {page}/{nb_pages} : {url}")
        time.sleep(DELAY_PAGES)
        html = fetch(url)
        if not html:
            break
        for m in LISTING_RE.finditer(html):
            lid = m.group(2)
            if lid not in results:
                results[lid] = absolute_listing_url(m.group(1))

    if total and len(results) < total:
        stats["complete"] = False
        print(
            "    [AVERTISSEMENT] Résultats incomplets : "
            f"{len(results)}/{total} annonces uniques collectées."
        )

    stats["collected"] = len(results)
    return (results, stats) if return_stats else results

# ---------------------------------------------------------------------------
# EXTRACTION D'UNE FICHE
# ---------------------------------------------------------------------------

def decode_html(text):
    text = html_module.unescape(text)
    return re.sub(r'\s+', ' ', text).strip()


def clean_number(text):
    if not text:
        return "Non indiqué"
    digits = re.sub(r'[^\d]', '', str(text))
    return int(digits) if digits else "Non indiqué"


def normalize_label(text):
    text = decode_html(re.sub(r'<[^>]+>', ' ', text))
    text = text.replace("’", "'").lower()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return re.sub(r"[^a-z0-9']+", " ", text).strip()


def address_key(adresse):
    if not adresse or adresse == "Non indiqué":
        return ""

    text = decode_html(str(adresse)).lower()
    text = text.replace("’", "'")
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = re.sub(r"[–—−]", "-", text)
    text = re.sub(r"\s*-\s*", "-", text)
    text = re.sub(r"[^a-z0-9#'()/-]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def row_address_key(row):
    key = address_key(row.get("Adresse", ""))
    if key:
        return key
    lid = row.get("_id", "")
    return f"id:{lid}" if lid else ""


def normalize_reference(ref):
    normalized = {}
    for key, value in ref.items():
        if isinstance(value, dict):
            normalized_key = address_key(value.get("adresse", "")) or key
            normalized[normalized_key] = value
        else:
            normalized[key] = value
    return normalized


def carac_value(html, label):
    expected = normalize_label(label)
    pat = re.compile(
        r'<div class="carac-title"[^>]*>(.*?)</div>\s*'
        r'<div class="carac-value"[^>]*>\s*<span[^>]*>(.*?)</span>',
        re.DOTALL | re.IGNORECASE
    )

    for title, value in pat.findall(html):
        if normalize_label(title) == expected:
            return decode_html(value)

    return "Non indiqué"


def financial_total(html, section_label):
    sec_pat = re.compile(
        section_label + r'.*?financial-details-table-total[^>]*>'
        r'.*?<td[^>]*class="font-weight-bold text-right"[^>]*>([\d\xA0\s,]+\$?)</td>',
        re.DOTALL | re.IGNORECASE
    )
    m = sec_pat.search(html)
    if m:
        return clean_number(m.group(1))

    label_idx = html.lower().find(section_label.lower())
    if label_idx == -1:
        return "Non indiqué"
    chunk   = html[label_idx: label_idx + 4000]
    total_m = re.search(
        r'font-weight-bold text-right"[^>]*>([\d\xA0\s,]+\$?)</td>',
        chunk, re.IGNORECASE
    )
    return clean_number(total_m.group(1)) if total_m else "Non indiqué"


def parse_unit_count(text):
    if not text or text == "Non indiqué":
        return "Non indiqué"
    m = re.search(r'(\d+)', str(text))
    return int(m.group(1)) if m else "Non indiqué"


def unit_count_from_url(url):
    type_counts = {
        "duplex": 2,
        "triplex": 3,
        "quadruplex": 4,
        "quintuplex": 5,
    }
    m = re.search(r'/fr/([^/~]+)~a-vendre~', url, re.IGNORECASE)
    if not m:
        return "Non indiqué"
    return type_counts.get(m.group(1).lower(), "Non indiqué")


def unit_count_from_mix(unit_mix):
    if not unit_mix or unit_mix == "Non indiqué":
        return "Non indiqué"

    parts = [part.strip() for part in str(unit_mix).split(",") if part.strip()]
    if not parts:
        return "Non indiqué"

    total = 0
    for part in parts:
        m = re.match(r'^(\d+)\s*x\b', part, re.IGNORECASE)
        if not m:
            return "Non indiqué"
        total += int(m.group(1))

    return total if total else "Non indiqué"


def resolve_unit_count(url, raw_count, unit_mix):
    for candidate in (
        parse_unit_count(raw_count),
        unit_count_from_url(url),
        unit_count_from_mix(unit_mix),
    ):
        if candidate != "Non indiqué":
            return candidate
    return "Non indiqué"


def extract_listing(url, ville_nom):
    html = fetch(url)
    if not html:
        return None

    data = {"Ville": ville_nom, "URL": url}

    addr_m = re.search(r'<h2\s+itemprop="address"[^>]*>\s*(.*?)\s*</h2>',
                       html, re.DOTALL | re.IGNORECASE)
    data["Adresse"] = decode_html(addr_m.group(1)) if addr_m else "Non indiqué"

    prix_m      = re.search(r'<meta\s+itemprop="price"\s+content="(\d+)"', html, re.IGNORECASE)
    data["Prix"] = int(prix_m.group(1)) if prix_m else "Non indiqué"

    data["Unités résidentielles"] = carac_value(html, "Unités résidentielles")
    nb_raw = carac_value(html, "Nombre d'unités")
    data["Nombre d'unités"] = resolve_unit_count(
        url,
        nb_raw,
        data["Unités résidentielles"],
    )

    annee_raw = carac_value(html, "Année de construction")
    if annee_raw != "Non indiqué":
        annee_m = re.search(r'(\d{4})', annee_raw)
        yr      = int(annee_m.group(1)) if annee_m else None
        data["Année de construction"] = yr if yr and 1800 <= yr <= 2030 else "Non indiqué"
    else:
        data["Année de construction"] = "Non indiqué"

    surf_raw                   = carac_value(html, "Superficie du terrain")
    data["Superficie terrain"] = clean_number(surf_raw)

    rev_raw                        = carac_value(html, "Revenus bruts potentiels")
    data["Revenu brut potentiel"]  = clean_number(rev_raw)

    data["Évaluation municipale"] = financial_total(html, "valuation municipale")
    data["Taxes"]                 = financial_total(html, "Taxes")
    data["Dépenses"]              = financial_total(html, "penses")

    # Stocker l'ID Centris dans la row pour la comparaison
    id_m        = re.search(r'/(\d{7,})$', url.rstrip('/'))
    data["_id"] = id_m.group(1) if id_m else url

    return data

# ---------------------------------------------------------------------------
# RÉFÉRENCE (JSON inter-runs)
# ---------------------------------------------------------------------------

def load_reference(path):
    """Charge le JSON de référence de la veille. Retourne {} si absent (premier run)."""
    if not os.path.exists(path):
        print("[RÉFÉRENCE] Aucun fichier de référence trouvé — premier run.")
        return {}
    try:
        with open(path, "r", encoding="utf-8") as f:
            ref = json.load(f)
        normalized = normalize_reference(ref)
        print(f"[RÉFÉRENCE] {len(normalized)} annonces chargées depuis la veille.")
        return normalized
    except Exception as e:
        print(f"[RÉFÉRENCE] Erreur de lecture : {e} — on repart de zéro.")
        return {}


def save_reference(rows, path):
    """Sauvegarde {address_key: {adresse, prix, ville, listing_id}} pour demain."""
    ref = {}
    for row in rows:
        key = row_address_key(row)
        lid = row.get("_id")
        if key:
            ref[key] = {
                "prix":    row.get("Prix", "Non indiqué"),
                "ville":   row.get("Ville", ""),
                "adresse": row.get("Adresse", ""),
                "listing_id": lid,
            }
    with open(path, "w", encoding="utf-8") as f:
        json.dump(ref, f, ensure_ascii=False, indent=2)
    print(f"[RÉFÉRENCE] {len(ref)} annonces sauvegardées → {path}")

# ---------------------------------------------------------------------------
# DÉTECTION DES CHANGEMENTS
# ---------------------------------------------------------------------------

def detect_changements(rows, ref):
    """
    Compare les annonces d'aujourd'hui avec la référence de la veille.

    Retourne :
        nouveaux_ids  : set — clés adresse absentes de ref (nouvelles annonces)
        retires       : list de dicts — adresses présentes hier, absentes aujourd'hui
        prix_changes  : dict {address_key: ancien_prix} — même adresse, prix différent
    """
    rows_by_key = {
        row_address_key(row): row
        for row in rows
        if row_address_key(row)
    }
    ids_aujourdhui = set(rows_by_key.keys())
    ids_hier       = set(ref.keys())

    nouveaux_ids = ids_aujourdhui - ids_hier
    retires_ids  = ids_hier - ids_aujourdhui

    retires = [
        {
            "id":         key,
            "adresse":    ref[key].get("adresse", "Adresse inconnue"),
            "ville":      ref[key].get("ville",   ""),
            "listing_id": ref[key].get("listing_id", ""),
        }
        for key in retires_ids
    ]

    prix_changes = {}
    for key, row in rows_by_key.items():
        if key in ref:
            ancien_prix   = ref[key].get("prix", "Non indiqué")
            nouveau_prix  = row.get("Prix", "Non indiqué")
            if (
                ancien_prix != "Non indiqué"
                and nouveau_prix != "Non indiqué"
                and ancien_prix != nouveau_prix
            ):
                prix_changes[key] = ancien_prix

    return nouveaux_ids, retires, prix_changes


def is_reference_update_safe(
    rows,
    ref,
    retires,
    collection_complete=True,
    expected_listing_count=None,
):
    """
    Retourne False si l'extraction ressemble à un run partiel.

    Sans ce garde-fou, une page Centris manquée aujourd'hui efface des annonces
    de la référence, puis ces mêmes annonces reviennent demain comme faux
    "nouveaux" plex.
    """
    if not collection_complete:
        print("[SÉCURITÉ] Extraction suspecte : collecte incomplète des résultats.")
        return False

    if expected_listing_count is not None and len(rows) < expected_listing_count:
        print(
            "[SÉCURITÉ] Extraction suspecte : "
            f"{len(rows)}/{expected_listing_count} fiches extraites."
        )
        return False

    previous_count = len(ref)
    current_count  = len(rows)
    removed_count  = len(retires)

    if previous_count == 0:
        return True

    active_ratio  = current_count / previous_count
    removed_ratio = removed_count / previous_count

    if active_ratio < MIN_ACTIVE_RATIO_FOR_SAFE_RUN:
        print(
            "[SÉCURITÉ] Extraction suspecte : "
            f"{current_count}/{previous_count} annonces actives "
            f"({active_ratio:.0%})."
        )
        return False

    if (
        removed_count > MAX_REMOVED_COUNT_FOR_SAFE_RUN
        or removed_ratio > MAX_REMOVED_RATIO_FOR_SAFE_RUN
    ):
        print(
            "[SÉCURITÉ] Trop d'annonces retirées en un run : "
            f"{removed_count}/{previous_count} ({removed_ratio:.0%})."
        )
        return False

    return True

# ---------------------------------------------------------------------------
# EXCEL
# ---------------------------------------------------------------------------

COLONNES = [
    "Ville",
    "Adresse",
    "Prix",
    "Nombre d'unités",
    "Année de construction",
    "Superficie terrain",
    "Unités résidentielles",
    "Revenu brut potentiel",
    "Évaluation municipale",
    "Taxes",
    "Dépenses",
    "URL",
]

COL_WIDTHS = {
    "Ville": 24,
    "Adresse": 46,
    "Prix": 14,
    "Nombre d'unités": 16,
    "Année de construction": 20,
    "Superficie terrain": 18,
    "Unités résidentielles": 32,
    "Revenu brut potentiel": 22,
    "Évaluation municipale": 22,
    "Taxes": 12,
    "Dépenses": 14,
    "URL": 65,
}

PRIX_COL_IDX = COLONNES.index("Prix") + 1   # index 1-based dans Excel


def save_excel(rows, path, nouveaux_ids=None, prix_changes=None):
    """
    Génère le fichier Excel.
    - Ligne verte   (#C6EFCE) si l'adresse est dans nouveaux_ids
    - Ligne jaune   (#FFEB9C) + cellule Prix orange (#FF6600 bold) si l'adresse est dans prix_changes
    """
    nouveaux_ids = nouveaux_ids or set()
    prix_changes  = prix_changes  or {}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Annonces admissibles"

    # En-tête
    h_font  = Font(bold=True, color="FFFFFF")
    h_fill  = PatternFill("solid", fgColor="1F4E79")
    h_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, col_name in enumerate(COLONNES, start=1):
        cell           = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font      = h_font
        cell.fill      = h_fill
        cell.alignment = h_align

    ws.row_dimensions[1].height = 30

    # Fills de highlighting
    fill_nouveau    = PatternFill("solid", fgColor=CLR_NOUVEAU)
    fill_prix_ligne = PatternFill("solid", fgColor=CLR_PRIX_LIGNE)
    fill_prix_cell  = PatternFill("solid", fgColor=CLR_PRIX_CELL)

    # Données
    for row_idx, row in enumerate(rows, start=2):
        key = row_address_key(row)

        # Déterminer le style de ligne
        if key in nouveaux_ids:
            row_fill = fill_nouveau
        elif key in prix_changes:
            row_fill = fill_prix_ligne
        else:
            row_fill = None

        for col_idx, col_name in enumerate(COLONNES, start=1):
            val  = row.get(col_name, "Non indiqué")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = Alignment(vertical="center")

            if row_fill:
                cell.fill = row_fill

            # Cellule Prix orange + bold si changement de prix
            if key in prix_changes and col_idx == PRIX_COL_IDX:
                cell.fill = fill_prix_cell
                cell.font = Font(bold=True, color="FFFFFF")

    # Largeurs de colonnes
    for col_idx, col_name in enumerate(COLONNES, start=1):
        letter = openpyxl.utils.get_column_letter(col_idx)
        ws.column_dimensions[letter].width = COL_WIDTHS.get(col_name, 20)

    ws.freeze_panes = "A2"

    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb.save(path)
    print(f"Fichier Excel sauvegardé : {path}")

# ---------------------------------------------------------------------------
# EMAIL
# ---------------------------------------------------------------------------

def _fmt_prix(prix):
    """Formate un prix entier en '419 900 $' ou retourne tel quel si non numérique."""
    try:
        return f"{int(prix):,}".replace(",", " ") + " $"
    except (ValueError, TypeError):
        return str(prix)


def send_email(excel_path, nb_total, nouveaux, retires, prix_changes, rows_dict):
    """
    Envoie l'email uniquement si des changements ont été détectés.
    rows_dict est conservé pour compatibilité avec l'appelant.
    """
    if not GMAIL_USER or not GMAIL_APP_PASSWORD or not EMAIL_DEST:
        print("[EMAIL] Variables d'environnement manquantes — email non envoyé.")
        return

    nb_nouveaux = len(nouveaux)
    nb_retires  = len(retires)
    nb_prix     = len(prix_changes)

    if nb_nouveaux + nb_retires + nb_prix == 0:
        print("[EMAIL] Aucun changement détecté — email non envoyé.")
        return

    today   = datetime.date.today().strftime("%d %B %Y")
    subject = (
        f"Centris — Changements détectés — {today} "
        f"(+{nb_nouveaux} / -{nb_retires} / ~{nb_prix} prix)"
    )

    # Corps de l'email: résumé seulement, les détails restent dans l'Excel.
    lignes = [
        f"Bonjour,",
        f"",
        f"Des modifications ont été détectées sur Centris ce matin ({today}).",
        f"",
        f"Résumé des changements :",
        f"  Nouvelles annonces  : {nb_nouveaux}",
        f"  Annonces retirées   : {nb_retires}",
        f"  Changements de prix : {nb_prix}",
        f"",
        f"Total actuel : {nb_total} annonces",
        f"",
        f"Le fichier Excel complet est joint (lignes colorées selon les changements).",
        f"  Vert   = nouvelle annonce",
        f"  Jaune  = changement de prix  (cellule Prix en orange)",
        f"",
        f"---",
        f"Généré automatiquement par le script Centris.",
    ]

    body = "\n".join(lignes)

    msg            = MIMEMultipart()
    msg["From"]    = GMAIL_USER
    msg["To"]      = EMAIL_DEST
    msg["Subject"] = subject
    msg.attach(MIMEText(body, "plain", "utf-8"))

    # Pièce jointe Excel
    filename = os.path.basename(excel_path)
    with open(excel_path, "rb") as f:
        part = MIMEBase(
            "application",
            "vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        part.set_payload(f.read())
    encoders.encode_base64(part)
    part.add_header("Content-Disposition", f'attachment; filename="{filename}"')
    msg.attach(part)

    try:
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
            server.login(GMAIL_USER, GMAIL_APP_PASSWORD)
            server.sendmail(GMAIL_USER, EMAIL_DEST, msg.as_string())
        print(f"[EMAIL] Envoyé avec succès à {EMAIL_DEST}")
    except Exception as e:
        print(f"[EMAIL] Erreur lors de l'envoi : {e}")

# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------

def main(reset_reference=False):
    print("=" * 65)
    print("EXTRACTION CENTRIS — PLEX À VENDRE — ~50 KM AUTOUR DE SHERBROOKE")
    print("=" * 65)

    # Étape 0 : charger la référence de la veille
    print("\n[ÉTAPE 0] Chargement de la référence de la veille...")
    ref = load_reference(REFERENCE_PATH)
    if reset_reference:
        print("[RÉFÉRENCE] Réinitialisation demandée : baseline reconstruite sans alerte.")
        ref = {}
    baseline_only = not ref

    # Étape 1 : collecter tous les liens
    print("\n[ÉTAPE 1] Collecte des liens d'annonces par ville...")
    all_listings = {}   # listing_id → (url, ville_nom)
    incomplete_villes = []

    for ville_nom, ville_slug in VILLES:
        print(f"\n  {ville_nom}")
        ville_res, ville_stats = get_listing_urls_for_ville(
            ville_nom,
            ville_slug,
            return_stats=True,
        )
        if not ville_stats["complete"]:
            incomplete_villes.append(
                f"{ville_nom} ({ville_stats['collected']}/{ville_stats['expected']})"
            )
        for lid, url in ville_res.items():
            if lid not in all_listings:
                all_listings[lid] = (url, ville_nom)
        time.sleep(DELAY_PAGES)

    total = len(all_listings)
    print(f"\n  Total annonces uniques : {total}")

    if total == 0:
        print("\nAucune annonce trouvée. Vérifiez la connectivité.")
        return

    # Étape 2 : extraire chaque fiche
    print("\n[ÉTAPE 2] Extraction des données de chaque fiche...")
    rows = []

    for idx, (lid, (url, ville_nom)) in enumerate(all_listings.items(), start=1):
        print(f"  [{idx:3d}/{total}] {ville_nom} — {url}")
        data = extract_listing(url, ville_nom)
        if data:
            rows.append(data)
        time.sleep(DELAY_FICHES)

    print(f"\n  Fiches extraites avec succès : {len(rows)}")
    if incomplete_villes:
        print(
            "\n  [AVERTISSEMENT] Villes incomplètes : "
            + ", ".join(incomplete_villes)
        )

    # Étape 3 : détecter les changements
    print("\n[ÉTAPE 3] Détection des changements...")
    nouveaux_ids, retires, prix_changes = detect_changements(rows, ref)

    if baseline_only:
        print("  Aucune référence exploitable : création d'une baseline, sans faux nouveaux.")
        nouveaux_ids = set()
        retires      = []
        prix_changes = {}

    print(f"  Nouvelles annonces  : {len(nouveaux_ids)}")
    print(f"  Annonces retirées   : {len(retires)}")
    print(f"  Changements de prix : {len(prix_changes)}")

    safe_reference_update = is_reference_update_safe(
        rows,
        ref,
        retires,
        collection_complete=not incomplete_villes,
        expected_listing_count=total,
    )

    # Étape 4 : générer l'Excel (annonces actives uniquement — retirées exclues)
    print("\n[ÉTAPE 4] Génération du fichier Excel...")
    rows_actifs = rows
    save_excel(rows_actifs, OUTPUT_PATH, nouveaux_ids, prix_changes)

    # Étape 5 : sauvegarder la nouvelle référence
    print("\n[ÉTAPE 5] Sauvegarde de la référence pour demain...")
    if safe_reference_update:
        save_reference(rows_actifs, REFERENCE_PATH)
    else:
        print("[SÉCURITÉ] Référence conservée : le run actuel semble partiel.")

    # Étape 6 : envoyer l'email si changements détectés
    print("\n[ÉTAPE 6] Envoi de l'email (si changements)...")
    rows_dict = {r["_id"]: r for r in rows if r.get("_id")}
    if safe_reference_update:
        send_email(
            excel_path   = OUTPUT_PATH,
            nb_total     = len(rows_actifs),
            nouveaux     = nouveaux_ids,
            retires      = retires,
            prix_changes = prix_changes,
            rows_dict    = rows_dict,
        )
    else:
        print("[EMAIL] Email non envoyé : extraction suspecte, rapport non fiable.")

    print("\nTerminé.")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Extraction Centris plex Sherbrooke.")
    parser.add_argument(
        "--reset-reference",
        action="store_true",
        help="Reconstruit la référence sans envoyer d'alerte de changements.",
    )
    args = parser.parse_args()
    main(reset_reference=args.reset_reference or RESET_REFERENCE)
